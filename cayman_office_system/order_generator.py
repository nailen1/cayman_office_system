from shining_pebbles import get_today
from .dataset_constants import file_folder
from .market_information import get_ks_market_info, get_ks_equity_info
import pandas as pd
from openpyxl.styles import Alignment

COLUMNS_INSTRUCTION = [
    'Name', 'Account', 'Side', 'Broker', 'Symbol', 'Amount', 
    'Order Type', 'Limit', 'Instructions', 'CFD', 'Rationale'
]

MAPPING_SIDE = {
    'Buy': 'B',
    'Sell': 'S',
    'B': 'B',
    'S': 'S'
}  

default_settings = lambda symbol:{
    'Name': None,
    'Account': '715543911501',
    'Broker': 'Samsung',
    'Order Type': 'MKT',
    'Limit': None,
    'Instructions': 'VWAP',
    'CFD': None,
    'Rationale': MAPPING_RATIONALE.get(symbol, None)
}

MAPPING_RATIONALE = {
    '007340 KS': "Currently trading at a PER of less than 3, indicating extreme undervaluation. The order backlog is steadily increasing, and favorable conditions (costs/FX) suggest consistent improvement in performance over the next 5 years. A cumulative profit increase of over 100% is possibled. In particular, IPO of the core subsidiary is expected within the next year. The target market cap presented by underwriters is at least five times the current market cap of the parent company, indicating a significant revaluation event is likely within the next year.",
    '282690 KS': "The company is set to merge with DN Automotive soon. Based on the merger valuation, the company's stock is currently trading at a discount compared to DN Automotive's stock, presenting an arbitrage opportunity.",
    '003030 KS': "The company is trading at a PER of less than 3, indicating extreme undervaluation. The market’s outlook on the future of the steel pipe industry is overly pessimistic. The current stock price does not yet reflect the upcoming operation of the UK offshore wind tower substructure plant, which is set to begin by the end of this year. The company is transitioning from a focus on fossil energy-related products to renewable energy products.",
    '271560 KS': "The company has been overlooked by the market due to its acquisition of a biotech firm focused on new drug development, which is unrelated to its traditional confectionery business. It is currently the most undervalued stock in the food and beverage sector. However, starting in the second half of the year, the company's core business is expected to see significant performance improvements, and the acquisition of the biotech company will gradually prove to have been a sound strategic move.",
    '000660 KS': "With the recent drop, PER has fallen below 5x, creating an attractive valuation opportunity. Concerns about AI investment are expected to ease after this fall. We have plans to engage to prevent profits and cash from being transferred to SK Holdings and its affiliates."
}

def remove_duplicate_dicts(dict_list):
    seen = set()
    unique_dicts = []
    
    for d in dict_list:
        dict_tuple = tuple(sorted(d.items()))
        if dict_tuple not in seen:
            seen.add(dict_tuple)
            unique_dicts.append(d)
    
    return unique_dicts


class OrderInstruction:
    def __init__(self, date=get_today()):
        self.date = date
        self.file_folder = file_folder['generate']
        self.frame = self.get_frame_of_instruction()
        self.orders = []
        
    def set_account(self, account=None):
        if account is None:
            account = '715543911501'
        self.account = account
        return account
    
    def set_default_sides(self, sides):
        self.default_sides = sides
        return sides

    def set_broker(self, broker=None):
        if broker is None:
            broker = 'Samsung'
        self.broker = broker
        return broker
    
    def set_order_type(self, order_type=None):
        if order_type is None:
            order_type = 'MKT'
        self.order_type = order_type
        return order_type
    
    def set_limit(self, limit=None):
        self.limit = limit
        return limit

    def set_instructions(self, instructions=None):
        if instructions is None:
            instructions = 'VWAP'
        self.instructions = instructions
        return instructions
    
    def set_cfd(self, cfd=None):
        self.cfd = cfd
        return cfd
    
    def set_order(self, order):
        orders = self.orders
        dct = {}
        for key in COLUMNS_INSTRUCTION:
            if key == 'Side':
                dct[key] = MAPPING_SIDE[order['Side']]
            elif key in order:
                dct[key] = order[key]
            else:
                dct[key] = default_settings(order['Symbol'])[key]
        if order['Limit'] != None:
            dct['Order Type'] = 'LMT'
        print(dct)
        orders.append(dct)
        self.orders = remove_duplicate_dicts(orders)
        return self
    
    def set_orders(self, orders):
        for order in orders:
            self.set_order(order)
        return self   
    
    def get_frame_of_instruction(self):
        df = pd.DataFrame(data=None, columns=COLUMNS_INSTRUCTION)
        self.frame = df
        return df

    def generate(self):
        df = pd.DataFrame(self.orders)
        df['Name'] = df['Symbol'].apply(lambda x: get_ks_market_info()['name'].loc[f'{x} Equity'])
        df = df.drop_duplicates()
        self.df = df
        self.tickers = list(df['Symbol'])
        return df            
    
    def show_current_equity_info(self):
        if not hasattr(self, 'tickers'):
            raise ValueError("No tickers exist. Generate orders first: OrderInstruction.generate()")
        tickers = self.tickers
        df = get_ks_equity_info(tickers)
        self.equities = df
        return df

    def export_as_excel(self):
        df = self.df
        df = df[COLUMNS_INSTRUCTION]
        file_folder = self.file_folder
        file_name = f'genearted_at{get_today("%Y%m%d%H")}-Life Asset trade instruction {self.date.replace("-", "")}.xlsx'
        file_path = os.path.join(file_folder, file_name)
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            worksheet = writer.sheets['Sheet1']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter 
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                    cell.alignment = Alignment(vertical='center')
                
                adjusted_width = (max_length + 4)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            last_column_index = worksheet.max_column            
            for col in worksheet.iter_cols(min_col=last_column_index, max_col=last_column_index):
                for cell in col:
                    cell.alignment = Alignment(wrapText=True, vertical='center')
            
            worksheet.column_dimensions[worksheet.cell(1, last_column_index).column_letter].width = 50 * 1.2
            
            for row in worksheet.iter_rows():
                max_font_size = 11 
                for cell in row:
                    try:
                        if cell.font.sz:  
                            max_font_size = max(max_font_size, cell.font.sz)
                    except:
                        pass
                worksheet.row_dimensions[cell.row].height = max_font_size * 3

        print(f"DataFrame has been saved to {file_path}")
        return None



    def show_current_price(self):
        df = self.df
        df['Price_last'] = df['Symbol'].apply(lambda x: get_ks_market_info()['price_last'].loc[f'{x} Equity'])
        df['Evaluation_krw'] = df['Price_last'] * df['Amount']
        cols_to_keep = ['Name', 'Symbol', 'Amount', 'Price_last', 'Evaluation_krw']
        df = df[cols_to_keep]
        self.current_price = df
        return df        